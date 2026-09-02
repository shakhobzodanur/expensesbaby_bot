# -*- coding: utf-8 -*-
"""Generate Excel export of user's transaction history."""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from lang import cat_name


def generate_excel(entries: list, currency: str, lang: str, balance: float) -> io.BytesIO:
    """
    entries: list of dicts with amount, type, category, created_at (newest first)
    Returns an in-memory xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = {
        "en": ["Date", "Time", "Type", "Category", "Amount", "Currency"],
        "ru": ["Дата", "Время", "Тип", "Категория", "Сумма", "Валюта"],
        "uz": ["Sana", "Vaqt", "Turi", "Kategoriya", "Miqdor", "Valyuta"],
    }.get(lang, ["Date", "Time", "Type", "Category", "Amount", "Currency"])

    type_label = {
        "en": {"income": "Income", "expense": "Expense"},
        "ru": {"income": "Доход",  "expense": "Расход"},
        "uz": {"income": "Daromad","expense": "Xarajat"},
    }.get(lang, {"income": "Income", "expense": "Expense"})

    # header row
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # data rows
    income_fill  = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    expense_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    row_idx = 2
    total_income = 0.0
    total_expense = 0.0

    for e in entries:
        dt = datetime.strptime(e["created_at"], "%Y-%m-%d %H:%M:%S")
        date_str = dt.strftime("%Y-%m-%d")
        time_str = dt.strftime("%H:%M")
        is_income = e["type"] == "income"
        amount = e["amount"] if is_income else -e["amount"]
        cat = cat_name(e["category"], lang) if e.get("category") else "—"

        ws.cell(row=row_idx, column=1, value=date_str)
        ws.cell(row=row_idx, column=2, value=time_str)
        ws.cell(row=row_idx, column=3, value=type_label["income" if is_income else "expense"])
        ws.cell(row=row_idx, column=4, value=cat)
        amount_cell = ws.cell(row=row_idx, column=5, value=amount)
        amount_cell.number_format = "#,##0"
        ws.cell(row=row_idx, column=6, value=currency)

        fill = income_fill if is_income else expense_fill
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).fill = fill

        if is_income:
            total_income += e["amount"]
        else:
            total_expense += e["amount"]

        row_idx += 1

    # summary rows
    row_idx += 1
    summary_label = {
        "en": ["Total Income", "Total Expense", "Balance"],
        "ru": ["Общий доход", "Общий расход", "Баланс"],
        "uz": ["Jami daromad", "Jami xarajat", "Balans"],
    }.get(lang, ["Total Income", "Total Expense", "Balance"])

    bold = Font(bold=True, size=11)
    for i, (label, value) in enumerate(zip(
        summary_label, [total_income, -total_expense, balance]
    )):
        r = row_idx + i
        ws.cell(row=r, column=4, value=label).font = bold
        c = ws.cell(row=r, column=5, value=value)
        c.font = bold
        c.number_format = "#,##0"

    # column widths
    widths = [12, 8, 10, 14, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
